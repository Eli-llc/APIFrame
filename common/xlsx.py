from copy import deepcopy
from openpyxl import load_workbook
import os

from common.log import logger
from common.error import NotFoundCaseId, FileContentError, ConfKeyNotFound, FileNotFoundError2
from common.parseyaml import config


logger = logger()


class XLSX:

    def __init__(self, name=None):
        # Handle data file absolute name
        if name:
            # If provided file name
            self.name = name
        else:
            try:
                # Read data file from config
                self.name = config.get_conf("file").get("case_file")
            except ConfKeyNotFound:
                # Use default file path
                self.name = os.path.abspath("./data/cmdb_data.xlsx")
        if not os.path.exists(self.name):
            message = "Data file not exist!"
            raise FileNotFoundError2(message)
        # Read contents from data file
        self.workbook = load_workbook(filename=self.name)
        self.sheets = self.workbook.sheetnames
        # Only use the 1st sheet.
        if len(self.sheets) > 1:
            logger.warn("The excel file have more than one sheet, now we use the first sheet.")

        self.content = self.workbook[self.sheets[0]]
        self.case_content = {}
        self.ids_content = {}

    def _case_content(self):
        # Get all rows by generator
        rows = self.content.rows
        # Get names of every column
        head_row = next(rows)
        heads = [x.value for x in head_row]
        # Store every case content to content_list
        content_list = []
        # Add every case content to content_list
        for case in rows:
            case_dict = {}
            current_case = [x.value for x in case]
            # Set value to every column's name as dict
            for head, value in zip(heads, current_case):
                case_dict[head] = value
            content_list.append(case_dict)
        return content_list

    def get_data_by_id(self, id_: (str, int)) -> dict:
        """
        You can get data by id of every case
        :param id_:
        :return: dict
        {
            "1": {
                'id': 1,
                'section': '初始化',
                'case_id': 16761,
                'case_name': '清空数据'
                ...
                }
        }
        """
        id_ = str(id_)
        # If ids_content is not empty, it already contains all cases
        if self.ids_content:
            return self.ids_content.get(id_)

        # Add all case to ids_content
        field_name = "id"
        case_content = self._case_content()
        for content in case_content:
            id_name = content.get(field_name, -1)
            self.ids_content[str(id_name)] = content
        return self.ids_content.get(id_)

    def get_data_by_case_id(self, column_key: (str, int), column="case_id") -> list:
        """
        Get a group data by case_id
        :param column: The column's name used to sort data
        :param column_key: get the column_key of data
        :return: list like below
        [
          {
            'id': 10,
            'section': '上传训练',
            'case_id': 2862
          },
          {
            'id': 11,
            'section': '上传训练',
            'case_id': 2862
          },
          {
            'id': 12,
            'section': '上传训练',
            'case_id': 2862
          }
        ]
        """
        column_key = str(column_key)
        if self.case_content:
            return self.case_content.get(column_key)
        case_content: list = self._case_content()
        for content in case_content:
            id_name = content.get(column, -1)
            case_id_str = str(id_name)
            # If self.case_content already has case_id_str, append case to it, otherwise, set to list first
            if case_id_str not in self.case_content:
                self.case_content[case_id_str] = []
            self.case_content[case_id_str].append(content)
        return self.case_content.get(column_key)


if __name__ == "__main__":
    # file_name = "/Users/eoitek/PycharmProjects/Alert2.3/data/alert2.3_test_data.xlsx"
    xlsx = XLSX(None)
    data = xlsx.get_data_by_case_id(17777)[2]
